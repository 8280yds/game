# -*- coding: utf-8 -*-
from openpyxl.reader.excel import load_workbook
import glob
import sys
import os
import time
import msvcrt
import shutil
import sys
import struct
import zlib

#获取脚本文件的当前路径
def cur_file_dir():
     #获取脚本路径
     _path = sys.path[0]
     #判断为脚本文件还是py2exe编译后的文件，如果是脚本文件，则返回的是脚本的目录，
	 #如果是py2exe编译后的文件，则返回的是编译后的文件路径
     if os.path.isdir(_path):
         return _path
     elif os.path.isfile(_path):
         return os.path.dirname(_path)
		 
#将相对路径转化为绝对路径
def relatively_to_abs(relaPath):
	relaPath = relaPath.strip()
	count = relaPath.index('/')
	relaPath = relaPath[count:].replace("/", "\\");
	
	absPath = path
	for i in range(1, count):
		absPath = absPath[:absPath.rindex('\\')]
	return absPath + relaPath
		 
#获取指定路径下的所有“xlsx”文件名
def getFileNames():
	fileNameArr = []
	for filePath in glob.glob(excelDirPath + "*" + extension):
		fileNameArr.append(os.path.basename(filePath)[:-len(extension)])
	return fileNameArr

#获取一个数据表的第一个页签的数据
def getSheetData(filename):
	workbook = load_workbook(excelDirPath + filename + extension)
	sheetnames = workbook.get_sheet_names()
	sheet = workbook.get_sheet_by_name(sheetnames[0])
	print   "load: ", filename + extension, "/", sheet.title
	if (sheet.get_highest_row()<2 or sheet.get_highest_column()<1) :
		print u"当前数据表标题不完整，按任意键退出:", ord(msvcrt.getch())
		
	if packVO :
		creatVOBySheet(changeStr(filename), sheet)
	if packXml or packBytes or packTotleXML:
		return creatXmlBySheet(filename, sheet)
	return ''
	
#转化字符串 将ss_ss_ss格式转化成SsSsSs的格式
def changeStr(_str):
	_newStr = ""
	_strArr = _str.split('_')
	for _s in _strArr:
		_newStr += _s.capitalize()
	return _newStr

#根据sheet创建vo
def creatVOBySheet(_name, sheet):
	voName = _name + 'DBVO'
	#处理数据
	voStr1 = 'using Freamwork;\nusing System.Xml;\n\npublic class ' + voName + ' : DBVO\n{'
	voStr2 = "\n\tpublic override void xmlToVo(XmlNode node)\n\t{"
	voStr2 += "\n\t\tXmlElement xmlelement = (XmlElement)node;"
	
	cols = sheet.get_highest_column()
	for _col in range(1, cols+1): 
		cell = sheet.cell(row = 2, column = _col)
		if not cell.value :
			continue
		_strArr = str(cell.value).split(':')
		if len(_strArr)>=2 and len(_strArr[0])>0 and len(_strArr[1])>0 :
			_attr = _strArr[0]
			_type = _strArr[1]
			voStr1 += "\n\t/// <summary>"
			voStr1 += "\n\t/// " + sheet.cell(row = 1, column = _col).value
			voStr1 += "\n\t/// </summary>"
			voStr1 += "\n\tpublic " + _type +" "+ _attr + ";\n"
			
			if typeDic.has_key(_type):
				voStr2 += "\n\t\t" + _attr + " = " + typeDic[_type].replace("{0}", _attr);
			else:
				print u"类型出错或未能找到(”+_type+“),按任意键退出:", ord(msvcrt.getch())
				
	voStr1 += voStr2;
	voStr1 += "\n\t}"
	voStr1 += "\n}"
	
	#写入
	csFile = open(voDirPath + voName + ".cs", 'w')
	csFile.write(voStr1)
	csFile.close()
	
#根据sheet创建xml和bytes
def creatXmlBySheet(filename, sheet):
	xmlStr = '<' + filename + '>'
	xmlStr2 = xmlTitleStr + '\n' + xmlStr
	rows = sheet.get_highest_row()
	cols = sheet.get_highest_column()
	hasDataArr = [False] * (rows + 1)
	xmlStrArr = ['<item'] * (rows + 1)
	
	for _col in range(1, cols+1) : 
		cell = sheet.cell(row = 2, column = _col)
		if not cell.value :
			continue
		_strArr = str(cell.value).split(':')
		if len(_strArr)>=2 and len(_strArr[0])>0 and len(_strArr[1])>0 :
			_attr = _strArr[0]
			
			for _row in range(3, rows+1) :
				cell = sheet.cell(row = _row, column = _col)
				if str(cell.value).strip() == '' :
					addStr = ''
				else :
					addStr = str(cell.value)
					hasDataArr[_row] = True
				xmlStrArr[_row] += ' ' + _attr + '="' + addStr + '"'
	
	for _row in range(3, rows+1) :
		if hasDataArr[_row] :
			xmlStrArr[_row] += '/>'
			xmlStr += xmlStrArr[_row]
			xmlStr2 += '\n\t' + xmlStrArr[_row]
			
	xmlStr += '</' + filename + '>'
	xmlStr2 += '\n' + '</' + filename + '>'
	
	#写入
	if packXml :
		xmlFile = open(xmlDirPath + filename + ".xml", 'w')
		xmlFile.write(xmlStr2)
		xmlFile.close()
	return xmlStr

#配置数据
xmlTitleStr = '<?xml version="1.0" encoding="UTF-8"?>'
extension = '.xlsx'			#excel文件后缀
packVO = True				#是否需要打包DBVO.CS文件
packXml = True				#是否需要打包XML文件
packTotleXML = True			#是否需要打包XML总文档
packBytes = False			#是否需要打包二进制文件
deleteAllVoBeforePack = True		#是否在打包前删除VO文件夹
deleteAllXmlBeforePack = True		#是否在打包前删除XML文件夹
excelDirPath = './Excel/'	#excel所在的文件夹
voDirPath = '../DBVO/'		#打包后存放VO的文件夹（相对于本python打包工具的位置，最好是独立文件夹）
xmlDirPath = './XML/'		#打包后存放XML的文件夹（相对于本python打包工具的位置，最好是独立文件夹）
totleXMLPath = './'			#打包后存放XML总文档的位置（相对于本python打包工具的位置）
byteDirPath = './'			#打包后存放Bytes的位置（相对于本python打包工具的位置）

typeDic = {}
typeDic['int'] = 'int.Parse(xmlelement.GetAttribute("{0}"));'
typeDic['string'] = 'xmlelement.GetAttribute("{0}");'
typeDic['bool'] = 'bool.Parse(xmlelement.GetAttribute("{0}"));'
typeDic['float'] = 'float.Parse(xmlelement.GetAttribute("{0}"));'
typeDic['uint'] = 'uint.Parse(xmlelement.GetAttribute("{0}"));'
typeDic['long'] = 'long.Parse(xmlelement.GetAttribute("{0}"));'
typeDic['ulong'] = 'ulong.Parse(xmlelement.GetAttribute("{0}"));'
typeDic['short'] = 'short.Parse(xmlelement.GetAttribute("{0}"));'
typeDic['ushort'] = 'ushort.Parse(xmlelement.GetAttribute("{0}"));'

reload(sys)
sys.setdefaultencoding('utf-8')
	
startTime = time.time()
path = cur_file_dir()

excelDirPath = relatively_to_abs(excelDirPath)
voDirPath = relatively_to_abs(voDirPath)
xmlDirPath = relatively_to_abs(xmlDirPath)
totleXMLPath = relatively_to_abs(totleXMLPath)
byteDirPath = relatively_to_abs(byteDirPath)

fileNameArr = getFileNames()

#VO文件夹处理
if packVO :
	if deleteAllVoBeforePack and os.path.exists(voDirPath):
		shutil.rmtree(voDirPath);
	os.makedirs(voDirPath)

#XML文件夹处理
if packXml :
	if deleteAllXmlBeforePack and os.path.exists(xmlDirPath):
		shutil.rmtree(xmlDirPath);
	os.makedirs(xmlDirPath)

#执行打包
totalXmlStr = ''
for filename in fileNameArr:
	totalXmlStr += getSheetData(filename)
	
if packTotleXML :
	totalXmlStr = '<root>' + totalXmlStr + '</root>'
	#写入all.xml
	totalXmlFile = open(byteDirPath + "all.xml", 'w')
	totalXmlFile.write(totalXmlStr)
	totalXmlFile.close()
	
if packBytes :
	compress = zlib.compressobj(1)
	byteFile = open(byteDirPath + "all.zbin", 'wb')
	byteFile.write(compress.compress(totalXmlStr))
	byteFile.write(compress.flush())
	byteFile.close()
	
print u"已经全部打包成功，按任意键退出:", ord(msvcrt.getch())
